import datetime
import html as _html
import json
import logging
import os
import re
import shutil
from collections import defaultdict
from collections.abc import Callable
from math import isnan, log, sqrt
from typing import Any

from django.conf import settings
from django.db.models import Count, Max, Min, Q, QuerySet
from django.template.loader import render_to_string

from webapp.models import Channel, Message

import networkx as nx
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

_ISOLATED_GRID_DIVISIONS: int = 200

type GraphData = dict[str, list[dict[str, Any]]]


def build_graph_data(
    graph: nx.DiGraph,
    channel_dict: dict[str, Any],
    positions: dict[str, tuple[float, float]],
) -> GraphData:
    """Serialize graph nodes and edges into the output dict."""
    graph_data: GraphData = {"nodes": [], "edges": []}

    for node_id, node_data in graph.nodes(data=True):
        pos = positions.get(node_data["data"]["pk"])
        node_info: dict[str, Any] = {
            "id": node_id,
            "x": float(pos[0]) if pos is not None else 0.0,
            "y": float(pos[1]) if pos is not None else 0.0,
        }
        for key in (
            "label",
            "communities",
            "color",
            "pic",
            "url",
            "activity_period",
            "fans",
            "in_deg",
            "is_lost",
            "messages_count",
            "out_deg",
        ):
            node_info[key] = node_data["data"][key]
        graph_data["nodes"].append(node_info)

    for index, (source, target, edge_data) in enumerate(graph.edges(data=True)):
        graph_data["edges"].append(
            {
                "source": source,
                "target": target,
                "weight": edge_data.get("weight", 0),
                "color": edge_data.get("color", ""),
                "id": index,
            }
        )

    return graph_data


def apply_base_node_measures(
    graph_data: GraphData,
    graph: nx.DiGraph,
    channel_dict: dict[str, Any],
    start_date: datetime.date | None = None,
    end_date: datetime.date | None = None,
) -> list[tuple[str, str]]:
    """Populate degree, fans, message count, and activity period on each node."""
    measures_labels: list[tuple[str, str]] = [
        ("in_deg", "Inbound connections"),
        ("out_deg", "Outbound connections"),
        ("fans", "Users"),
        ("messages_count", "Messages"),
    ]

    channel_pks = [
        channel_dict[node["id"]]["channel"].pk for node in graph_data["nodes"] if channel_dict.get(node["id"])
    ]
    msg_q = Q(channel_id__in=channel_pks)
    if start_date:
        msg_q &= Q(date__date__gte=start_date)
    if end_date:
        msg_q &= Q(date__date__lte=end_date)
    message_counts: dict[int, int] = {
        item["channel_id"]: item["total"]
        for item in Message.objects.filter(msg_q).values("channel_id").annotate(total=Count("id"))
    }
    activity_bounds: dict[int, dict] = {
        item["channel_id"]: {"min_date": item["min_date"], "max_date": item["max_date"]}
        for item in Message.objects.filter(channel_id__in=channel_pks, date__isnull=False)
        .values("channel_id")
        .annotate(min_date=Min("date"), max_date=Max("date"))
    }

    now = datetime.datetime.now(datetime.timezone.utc)
    date_template = "%b %Y"
    for node in graph_data["nodes"]:
        channel_entry = channel_dict.get(node["id"])
        if channel_entry is None:
            continue
        channel = channel_entry["channel"]
        node["in_deg"] = graph.in_degree(node["id"], weight="weight")
        node["out_deg"] = graph.out_degree(node["id"], weight="weight")
        node["fans"] = channel.participants_count
        node["messages_count"] = message_counts.get(channel.pk, 0)
        node["label"] = channel.title
        agg = activity_bounds.get(channel.pk, {})
        first_date, last_date = agg.get("min_date"), agg.get("max_date")
        start_candidates = [d for d in (channel.date, first_date) if d is not None]
        end_candidates = [d for d in (channel.date, last_date) if d is not None]
        start = min(start_candidates) if start_candidates else None
        end = max(end_candidates) if end_candidates else None
        if start is None or end is None:
            node["activity_period"] = "Unknown"
            node["activity_start"] = ""
            node["activity_end"] = ""
        else:
            node["activity_period"] = (
                f"{start.strftime(date_template)} - {end.strftime(date_template)}"
                if end < now - datetime.timedelta(days=30)
                else f"{start.strftime(date_template)} - "
            )
            node["activity_start"] = start.strftime("%Y-%m")
            node["activity_end"] = end.strftime("%Y-%m")
    return measures_labels


def apply_pagerank(graph_data: GraphData, graph: nx.DiGraph) -> list[tuple[str, str]]:
    """Add PageRank score to each node."""
    key = "pagerank"
    pagerank_values: dict[str, float] = nx.pagerank(graph)
    for node in graph_data["nodes"]:
        if node["id"] in pagerank_values:
            node[key] = pagerank_values[node["id"]]
    return [(key, "PageRank")]


def apply_hits(graph_data: GraphData, graph: nx.DiGraph) -> list[tuple[str, str]]:
    """Add HITS hub and authority scores to each node."""
    try:
        hubs, authorities = nx.hits(graph)
    except nx.PowerIterationFailedConvergence:
        logger.warning("HITS failed to converge")
        return []
    for node in graph_data["nodes"]:
        node["hits_hub"] = hubs.get(node["id"], 0.0)
        node["hits_authority"] = authorities.get(node["id"], 0.0)
    return [("hits_hub", "HITS Hub"), ("hits_authority", "HITS Authority")]


def apply_betweenness_centrality(graph_data: GraphData, graph: nx.DiGraph) -> list[tuple[str, str]]:
    """Add betweenness centrality to each node."""
    key = "betweenness"
    values: dict[str, float] = nx.betweenness_centrality(graph, weight="weight")
    for node in graph_data["nodes"]:
        node[key] = values.get(node["id"], 0.0)
    return [(key, "Betweenness Centrality")]


def apply_in_degree_centrality(graph_data: GraphData, graph: nx.DiGraph) -> list[tuple[str, str]]:
    """Add normalized in-degree centrality to each node."""
    key = "in_degree_centrality"
    values: dict[str, float] = nx.in_degree_centrality(graph)
    for node in graph_data["nodes"]:
        node[key] = values.get(node["id"], 0.0)
    return [(key, "In-degree Centrality")]


def apply_out_degree_centrality(graph_data: GraphData, graph: nx.DiGraph) -> list[tuple[str, str]]:
    """Add normalized out-degree centrality to each node."""
    key = "out_degree_centrality"
    values: dict[str, float] = nx.out_degree_centrality(graph)
    for node in graph_data["nodes"]:
        node[key] = values.get(node["id"], 0.0)
    return [(key, "Out-degree Centrality")]


def apply_harmonic_centrality(graph_data: GraphData, graph: nx.DiGraph) -> list[tuple[str, str]]:
    """Add normalized harmonic centrality to each node."""
    key = "harmonic_centrality"
    n = graph.number_of_nodes()
    norm = (n - 1) if n > 1 else 1
    values: dict[str, float] = nx.harmonic_centrality(graph)
    for node in graph_data["nodes"]:
        node[key] = values.get(node["id"], 0.0) / norm
    return [(key, "Harmonic Centrality")]


def apply_katz_centrality(graph_data: GraphData, graph: nx.DiGraph) -> list[tuple[str, str]]:
    """Add Katz centrality to each node."""
    key = "katz_centrality"
    try:
        values: dict[str, float] = nx.katz_centrality(graph, weight="weight")
    except nx.PowerIterationFailedConvergence:
        logger.warning("Katz centrality failed to converge")
        return []
    for node in graph_data["nodes"]:
        node[key] = values.get(node["id"], 0.0)
    return [(key, "Katz Centrality")]


def apply_bridging_centrality(graph_data: GraphData, graph: nx.DiGraph, strategy_key: str) -> list[tuple[str, str]]:
    """Add bridging centrality (betweenness × neighbor-community Shannon entropy) to each node.

    For each node, the Shannon entropy is computed over the community distribution of its
    neighbours weighted by edge strength. Nodes that connect many distinct communities score
    high on entropy; multiplying by betweenness surfaces nodes that are both structurally
    central and community-diverse.
    """
    key = "bridging_centrality"

    betweenness: dict[str, float] = nx.betweenness_centrality(graph, weight="weight")

    community_map: dict[str, str] = {
        node_id: node_data["communities"][strategy_key]
        for node_id, node_data in graph.nodes(data="data")
        if node_data and strategy_key in (node_data.get("communities") or {})
    }

    for node in graph_data["nodes"]:
        node_id = node["id"]
        bt = betweenness.get(node_id, 0.0)

        community_weights: dict[str, float] = {}
        for pred in graph.predecessors(node_id):
            w = graph.edges[pred, node_id].get("weight", 1.0)
            c = community_map.get(pred)
            if c is not None:
                community_weights[c] = community_weights.get(c, 0.0) + w
        for succ in graph.successors(node_id):
            w = graph.edges[node_id, succ].get("weight", 1.0)
            c = community_map.get(succ)
            if c is not None:
                community_weights[c] = community_weights.get(c, 0.0) + w

        total = sum(community_weights.values())
        if total == 0.0 or len(community_weights) <= 1:
            entropy = 0.0
        else:
            entropy = -sum((w / total) * log(w / total) for w in community_weights.values())

        node[key] = bt * entropy

    return [(key, "Bridging Centrality")]


def apply_burt_constraint(graph_data: GraphData, graph: nx.DiGraph) -> list[tuple[str, str]]:
    """Add Burt's constraint to each node. Isolated nodes receive None (undefined)."""
    key = "burt_constraint"
    values: dict[str, float] = nx.constraint(graph)
    for node in graph_data["nodes"]:
        val = values.get(node["id"])
        node[key] = None if (val is None or isnan(val)) else round(val, 6)
    return [(key, "Burt's Constraint")]


def apply_amplification_factor(
    graph_data: GraphData,
    graph: nx.DiGraph,
    channel_dict: dict[str, Any],
    start_date: datetime.date | None = None,
    end_date: datetime.date | None = None,
) -> list[tuple[str, str]]:
    """Add amplification factor (forwards received / own message count) to each node."""
    key = "amplification_factor"

    channel_pks = [
        channel_dict[node["id"]]["channel"].pk for node in graph_data["nodes"] if channel_dict.get(node["id"])
    ]
    msg_q = Q(channel_id__in=channel_pks)
    if start_date:
        msg_q &= Q(date__date__gte=start_date)
    if end_date:
        msg_q &= Q(date__date__lte=end_date)
    message_counts: dict[int, int] = {
        item["channel_id"]: item["total"]
        for item in Message.objects.filter(msg_q).values("channel_id").annotate(total=Count("id"))
    }

    fwd_q = Q(forwarded_from_id__in=channel_pks, channel__organization__is_interesting=True)
    if start_date:
        fwd_q &= Q(date__date__gte=start_date)
    if end_date:
        fwd_q &= Q(date__date__lte=end_date)
    forwards_received: dict[int, int] = {
        item["forwarded_from_id"]: item["total"]
        for item in Message.objects.filter(fwd_q).values("forwarded_from_id").annotate(total=Count("id"))
    }

    for node in graph_data["nodes"]:
        channel_entry = channel_dict.get(node["id"])
        if channel_entry is None:
            continue
        pk = channel_entry["channel"].pk
        mc = message_counts.get(pk, 0)
        fr = forwards_received.get(pk, 0)
        node[key] = round(fr / mc, 4) if mc > 0 else 0.0

    return [(key, "Amplification Factor")]


def apply_content_originality(
    graph_data: GraphData,
    graph: nx.DiGraph,
    channel_dict: dict[str, Any],
    start_date: datetime.date | None = None,
    end_date: datetime.date | None = None,
) -> list[tuple[str, str]]:
    """Add content originality (1 − forwarded_messages / total_messages) to each node. None if no messages."""
    key = "content_originality"

    channel_pks = [
        channel_dict[node["id"]]["channel"].pk for node in graph_data["nodes"] if channel_dict.get(node["id"])
    ]
    msg_q = Q(channel_id__in=channel_pks)
    if start_date:
        msg_q &= Q(date__date__gte=start_date)
    if end_date:
        msg_q &= Q(date__date__lte=end_date)
    message_counts: dict[int, int] = {
        item["channel_id"]: item["total"]
        for item in Message.objects.filter(msg_q).values("channel_id").annotate(total=Count("id"))
    }
    fwd_q = msg_q & Q(forwarded_from__isnull=False)
    forwarded_counts: dict[int, int] = {
        item["channel_id"]: item["total"]
        for item in Message.objects.filter(fwd_q).values("channel_id").annotate(total=Count("id"))
    }

    for node in graph_data["nodes"]:
        channel_entry = channel_dict.get(node["id"])
        if channel_entry is None:
            continue
        pk = channel_entry["channel"].pk
        mc = message_counts.get(pk, 0)
        node[key] = round(1 - forwarded_counts.get(pk, 0) / mc, 4) if mc > 0 else None

    return [(key, "Content Originality")]


def find_main_component(graph: nx.DiGraph) -> set[str]:
    return max(nx.weakly_connected_components(graph), key=len)


def reposition_isolated_nodes(graph_data: GraphData, main_component: set[str]) -> None:
    """Move isolated nodes (outside the main component) into a grid near the main cluster."""
    main_nodes = [node for node in graph_data["nodes"] if node["id"] in main_component]
    isolated_nodes = [index for index, node in enumerate(graph_data["nodes"]) if node["id"] not in main_component]
    if not main_nodes:
        return
    max_x = max(node["x"] for node in main_nodes)
    min_x = min(node["x"] for node in main_nodes)
    max_y = max(node["y"] for node in main_nodes)
    d = abs(max_x - min_x) / _ISOLATED_GRID_DIVISIONS if max_x != min_x else 1.0
    col = int(sqrt(len(isolated_nodes))) + 1
    for i in range(col):
        for j in range(col):
            idx = i * col + j
            if len(isolated_nodes) > idx:
                graph_data["nodes"][isolated_nodes[idx]]["x"] = max_x - i * d
                graph_data["nodes"][isolated_nodes[idx]]["y"] = max_y - j * d


def ensure_graph_root(root_target: str) -> None:
    if os.path.isdir(root_target):
        for entry in os.scandir(root_target):
            if entry.is_dir(follow_symlinks=False):
                shutil.rmtree(entry.path)
            else:
                os.remove(entry.path)
    else:
        os.makedirs(root_target)
    try:
        shutil.copytree("webapp_engine/map", root_target, dirs_exist_ok=True)
    except OSError as e:
        logger.warning("Could not copy map template to %s: %s", root_target, e)


def _patch_html_file(path: str, seo: bool, project_title: str) -> None:
    """Patch the robots meta tag and title in a static HTML file in-place."""
    if not os.path.exists(path):
        return
    with open(path) as f:
        content = f.read()
    if seo:
        content = content.replace(
            '  <meta name="robots" content="noindex">',
            '  <meta name="robots" content="index, follow">',
        )
    if project_title:
        escaped = _html.escape(project_title)
        content = re.sub(r"<title>[^<]*</title>", f"<title>{escaped}</title>", content)
        content = re.sub(
            r'(<h4 class="modal-title" id="about_modalLabel">)[^<]*(</h4>)',
            rf"\g<1>{escaped}\g<2>",
            content,
        )
    with open(path, "w") as f:
        f.write(content)


def apply_robots_to_graph_html(root_target: str, seo: bool, project_title: str = "", include_3d: bool = False) -> None:
    """Patch the robots meta tag and title in the static graph HTML files after they are copied."""
    _patch_html_file(os.path.join(root_target, "graph.html"), seo, project_title)
    if include_3d:
        _patch_html_file(os.path.join(root_target, "graph3d.html"), seo, project_title)


def write_robots_txt(root_target: str, seo: bool) -> None:
    """Write a robots.txt that either allows or disallows all crawlers."""
    if seo:
        content = "User-agent: *\nAllow: /\n"
    else:
        content = "User-agent: *\nDisallow: /\n"
    with open(os.path.join(root_target, "robots.txt"), "w") as f:
        f.write(content)


def write_graph_files(
    graph_data: GraphData,
    communities_data: dict[str, Any],
    measures_labels: list[tuple[str, str]],
    channel_qs: "QuerySet[Channel]",
    graph_dir: str,
    include_positions: bool = True,
    positions_3d: dict | None = None,
) -> None:
    data_dir = os.path.join(graph_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    if include_positions:
        # channel_position.json — spatial layout + edges
        position_payload = {
            "nodes": [{"id": n["id"], "x": n["x"], "y": n["y"]} for n in graph_data["nodes"]],
            "edges": graph_data["edges"],
        }
        with open(os.path.join(data_dir, "channel_position.json"), "w") as f:
            f.write(json.dumps(position_payload))

    if positions_3d is not None:
        # channel_position_3d.json — 3D spatial layout + edges
        nodes_3d = []
        for n in graph_data["nodes"]:
            pos = positions_3d.get(n["id"])
            nodes_3d.append(
                {
                    "id": n["id"],
                    "x": float(pos[0]) if pos is not None else 0.0,
                    "y": float(pos[1]) if pos is not None else 0.0,
                    "z": float(pos[2]) if pos is not None else 0.0,
                }
            )
        position_3d_payload = {"nodes": nodes_3d, "edges": graph_data["edges"]}
        with open(os.path.join(data_dir, "channel_position_3d.json"), "w") as f:
            f.write(json.dumps(position_3d_payload))

    # channels.json — per-node metadata, computed measures, community assignments, measure labels
    node_keys: set[str] = {
        "id",
        "label",
        "color",
        "pic",
        "url",
        "activity_period",
        "fans",
        "in_deg",
        "is_lost",
        "messages_count",
        "out_deg",
        "activity_start",
        "activity_end",
    } | {k for k, _ in measures_labels}
    channels_payload: dict[str, Any] = {
        "nodes": [
            {**{k: n[k] for k in node_keys if k in n}, "communities": n.get("communities", {})}
            for n in graph_data["nodes"]
        ],
        "measures": measures_labels,
        "total_pages_count": channel_qs.count(),
    }
    with open(os.path.join(data_dir, "channels.json"), "w") as f:
        f.write(json.dumps(channels_payload))

    # communities.json — strategy group definitions (metrics rows added later by write_community_metrics_json)
    with open(os.path.join(data_dir, "communities.json"), "w") as f:
        f.write(json.dumps({"strategies": communities_data}))


def copy_channel_media(channel_qs: QuerySet[Channel], root_target: str) -> None:
    for username, telegram_id in channel_qs.values_list("username", "telegram_id"):
        channel_dir = username or str(telegram_id)
        src = os.path.join(settings.MEDIA_ROOT, "channels", channel_dir, "profile")
        dst = os.path.join(root_target, "channels", channel_dir, "profile")
        try:
            shutil.copytree(src, dst, dirs_exist_ok=True)
        except FileNotFoundError:
            pass
        except OSError as e:
            logger.warning("Could not copy media for channel %s: %s", channel_dir, e)


_BASE_MEASURE_KEYS: frozenset[str] = frozenset({"in_deg", "out_deg", "fans", "messages_count"})


def _heatmap_bg(val: float | int | None, col_min: float, col_max: float) -> str:
    """Subtle blue heatmap cell background: white (min) → #dceaf9 (max)."""
    if val is None or col_min >= col_max:
        return ""
    ratio = (val - col_min) / (col_max - col_min)
    r = round(255 - ratio * 35)
    g = round(255 - ratio * 21)
    b = round(255 - ratio * 6)
    return f"background-color:rgb({r},{g},{b})"


def _num_cell_dict(val: Any, key: str, hm_ranges: dict[str, tuple[float, float]]) -> dict:
    bg = _heatmap_bg(val, *hm_ranges[key]) if key in hm_ranges else ""
    return {
        "display": str(val) if val is not None else "",
        "css_class": "number",
        "sort_value": str(val) if val is not None else "",
        "style": bg,
        "link": "",
    }


def _float_cell_dict(val: Any, key: str, hm_ranges: dict[str, tuple[float, float]]) -> dict:
    bg = _heatmap_bg(val, *hm_ranges[key]) if key in hm_ranges else ""
    return {
        "display": f"{val:.4f}" if val is not None else "",
        "css_class": "number",
        "sort_value": str(val) if val is not None else "",
        "style": bg,
        "link": "",
    }


def write_table_html(
    graph_data: GraphData,
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    n = len(graph_data["nodes"])
    if seo:
        title = f"{project_title} | Channels" if project_title else "Channel network data"
        robots_meta = "index, follow"
    else:
        title = f"{project_title} | Channels" if project_title else "Channels"
        robots_meta = "noindex, nofollow"

    context = {
        "title": title,
        "robots_meta": robots_meta,
        "description": (
            f"Network data for {n} Telegram channels, "
            "including activity metrics, inbound and outbound connections, and community assignments."
        ),
    }
    content = render_to_string("network/channel_table.html", context)
    with open(output_filename, "w") as f:
        f.write(content)


def write_table_xlsx(
    graph_data: GraphData,
    measures_labels: list[tuple[str, str]],
    strategies: list[str],
    output_filename: str,
    project_title: str = "",
) -> None:
    extra = [(k, lbl) for k, lbl in measures_labels if k not in _BASE_MEASURE_KEYS]
    pagerank_col = next(((k, lbl) for k, lbl in extra if k == "pagerank"), None)
    other_extra = [(k, lbl) for k, lbl in extra if k != "pagerank"]
    nodes = sorted(graph_data["nodes"], key=lambda n: n.get("in_deg") or 0, reverse=True)

    wb = openpyxl.Workbook()
    if project_title:
        wb.properties.title = project_title
    ws = wb.active
    ws.title = "Channels"

    headers = ["Channel", "URL", "Users", "Messages", "Inbound", "Outbound"]
    if pagerank_col:
        headers.append(pagerank_col[1])
    headers += [lbl for _, lbl in other_extra]
    headers += [s.capitalize() for s in strategies]
    headers += ["Activity start", "Activity end"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for node in nodes:
        communities = node.get("communities") or {}
        row: list[Any] = [
            node.get("label") or node["id"],
            node.get("url") or "",
            node.get("fans"),
            node.get("messages_count"),
            node.get("in_deg"),
            node.get("out_deg"),
        ]
        if pagerank_col:
            row.append(node.get(pagerank_col[0]))
        for key, _ in other_extra:
            row.append(node.get(key))
        for s in strategies:
            row.append(communities.get(s, ""))
        row.append(node.get("activity_start") or "")
        row.append(node.get("activity_end") or "")
        ws.append(row)

    wb.save(output_filename)


def _network_summary(graph: nx.DiGraph) -> dict[str, Any]:
    """Compute structural metrics for the whole graph."""
    n = graph.number_of_nodes()
    e = graph.number_of_edges()
    density = nx.density(graph)
    try:
        reciprocity = nx.overall_reciprocity(graph) if e > 0 else 0.0
    except Exception:
        reciprocity = None
    try:
        avg_clustering = nx.average_clustering(graph)
    except Exception:
        avg_clustering = None
    avg_path_length = None
    diameter = None
    path_on_full = False
    wcc_count = None
    wcc_fraction = None
    scc_count = None
    scc_fraction = None
    if n >= 2:
        try:
            wccs = list(nx.weakly_connected_components(graph))
            largest_wcc = max(wccs, key=len)
            path_on_full = len(largest_wcc) == n
            wcc_count = len(wccs)
            wcc_fraction = len(largest_wcc) / n
            if len(largest_wcc) >= 2:
                ug = graph.subgraph(largest_wcc).to_undirected()
                avg_path_length = nx.average_shortest_path_length(ug)
                diameter = nx.diameter(ug)
        except Exception:
            pass
        try:
            sccs = list(nx.strongly_connected_components(graph))
            largest_scc = max(sccs, key=len)
            scc_count = len(sccs)
            scc_fraction = len(largest_scc) / n
        except Exception:
            pass
    assortativity: dict[str, float | None] = {
        "in_in": None,
        "in_out": None,
        "out_in": None,
        "out_out": None,
    }
    if e >= 2:
        try:
            in_deg = dict(graph.in_degree())
            out_deg = dict(graph.out_degree())
            src_in = np.array([in_deg[u] for u, v in graph.edges()], dtype=float)
            src_out = np.array([out_deg[u] for u, v in graph.edges()], dtype=float)
            tgt_in = np.array([in_deg[v] for u, v in graph.edges()], dtype=float)
            tgt_out = np.array([out_deg[v] for u, v in graph.edges()], dtype=float)
            for key, x, y in [
                ("in_in", src_in, tgt_in),
                ("in_out", src_in, tgt_out),
                ("out_in", src_out, tgt_in),
                ("out_out", src_out, tgt_out),
            ]:
                if x.std() > 0 and y.std() > 0:
                    assortativity[key] = float(np.corrcoef(x, y)[0, 1])
        except Exception:
            pass
    return {
        "n": n,
        "e": e,
        "density": density,
        "reciprocity": reciprocity,
        "avg_clustering": avg_clustering,
        "avg_path_length": avg_path_length,
        "diameter": diameter,
        "path_on_full": path_on_full,
        "wcc_count": wcc_count,
        "wcc_fraction": wcc_fraction,
        "scc_count": scc_count,
        "scc_fraction": scc_fraction,
        "assortativity": assortativity,
    }


def _subgraph_metrics(nodes_set: set[str], graph: nx.DiGraph) -> dict[str, Any]:
    """Compute structural metrics for a community defined by nodes_set."""
    subgraph = graph.subgraph(nodes_set)
    n = subgraph.number_of_nodes()
    internal_edges = subgraph.number_of_edges()
    total_deg = sum(graph.in_degree(nd) + graph.out_degree(nd) for nd in nodes_set)
    external_edges = total_deg - 2 * internal_edges
    density = nx.density(subgraph)
    try:
        reciprocity = nx.overall_reciprocity(subgraph) if internal_edges > 0 else 0.0
    except Exception:
        reciprocity = None
    try:
        avg_clustering = nx.average_clustering(subgraph)
    except Exception:
        avg_clustering = None
    avg_path_length = None
    diameter = None
    if n >= 2:
        try:
            wccs = list(nx.weakly_connected_components(subgraph))
            largest_wcc = max(wccs, key=len)
            if len(largest_wcc) >= 2:
                ug = subgraph.subgraph(largest_wcc).to_undirected()
                avg_path_length = nx.average_shortest_path_length(ug)
                diameter = nx.diameter(ug)
        except Exception:
            pass
    return {
        "internal_edges": internal_edges,
        "external_edges": external_edges,
        "density": density,
        "reciprocity": reciprocity,
        "avg_clustering": avg_clustering,
        "avg_path_length": avg_path_length,
        "diameter": diameter,
    }


def _freeman_centralization(values: list[float]) -> float | None:
    """Freeman (1978) graph centralization for a centrality measure.

    H = Σ_i (C_max - C_i) / [(n-1) · C_max]

    Returns None when the result is undefined (fewer than 2 nodes or C_max == 0).
    None entries in values are ignored.
    """
    clean = [v for v in values if v is not None]
    n = len(clean)
    if n < 2:
        return None
    c_max = max(clean)
    if c_max == 0:
        return None
    return sum(c_max - v for v in clean) / ((n - 1) * c_max)


type CommunityTableData = dict[str, Any]
# Structure:
# {
#   "network_summary": dict,          # from _network_summary() plus "centralizations"
#   "strategies": {
#     strategy_key: [                 # ordered as in communities_data
#       {"group": tuple, "node_count": int, "metrics": dict},
#       ...
#     ]
#   }
# }


def _network_content_metrics(
    channel_qs: QuerySet,
    start_date: datetime.date | None = None,
    end_date: datetime.date | None = None,
) -> dict[str, float | None]:
    """Compute network-wide content originality and amplification ratio from the DB."""
    channel_pks = list(channel_qs.values_list("pk", flat=True))
    msg_q = Q(channel_id__in=channel_pks)
    if start_date:
        msg_q &= Q(date__date__gte=start_date)
    if end_date:
        msg_q &= Q(date__date__lte=end_date)
    total = Message.objects.filter(msg_q).count()
    if total == 0:
        return {"network_originality": None, "network_amplification": None}
    forwarded_out = Message.objects.filter(msg_q & Q(forwarded_from__isnull=False)).count()
    fwd_in_q = Q(forwarded_from_id__in=channel_pks, channel__organization__is_interesting=True)
    if start_date:
        fwd_in_q &= Q(date__date__gte=start_date)
    if end_date:
        fwd_in_q &= Q(date__date__lte=end_date)
    forwards_received = Message.objects.filter(fwd_in_q).count()
    return {
        "network_originality": round(1 - forwarded_out / total, 4),
        "network_amplification": round(forwards_received / total, 4),
    }


def compute_community_metrics(
    graph_data: GraphData,
    communities_data: dict[str, Any],
    graph: nx.DiGraph,
    strategies: list[str],
    measures_labels: "list[tuple[str, str]] | None" = None,
    status_callback: "Callable[[str], None] | None" = None,
    channel_qs: "QuerySet | None" = None,
    start_date: datetime.date | None = None,
    end_date: datetime.date | None = None,
) -> CommunityTableData:
    """Pre-compute all structural metrics needed for community table outputs.

    ``measures_labels`` is the list of (node_key, display_label) pairs returned by the
    apply_* functions; when provided, Freeman centralization is computed for each measure.
    ``status_callback`` is called with a short label after each step completes
    so the caller can emit progress output between steps.
    ``channel_qs`` enables whole-network content originality and amplification ratio metrics.
    """
    network_summary = _network_summary(graph)
    centralizations: dict[str, tuple[float | None, str]] = {}
    if measures_labels:
        for key, label in measures_labels:
            values = [node[key] for node in graph_data["nodes"] if key in node]
            centralizations[key] = (_freeman_centralization(values), label)
    network_summary["centralizations"] = centralizations
    constraint_vals = [
        node["burt_constraint"] for node in graph_data["nodes"] if node.get("burt_constraint") is not None
    ]
    network_summary["mean_burt_constraint"] = sum(constraint_vals) / len(constraint_vals) if constraint_vals else None
    if channel_qs is not None:
        network_summary.update(_network_content_metrics(channel_qs, start_date, end_date))
    result: CommunityTableData = {"network_summary": network_summary, "strategies": {}}
    id_to_node: dict[str, dict] = {node["id"]: node for node in graph_data["nodes"]}
    if status_callback:
        status_callback("network")
    for strategy_key in strategies:
        strategy_data = communities_data.get(strategy_key)
        if not strategy_data:
            if status_callback:
                status_callback(strategy_key)
            continue
        label_to_nodes: dict[str, set[str]] = defaultdict(set)
        for node in graph_data["nodes"]:
            lbl = (node.get("communities") or {}).get(strategy_key)
            if lbl:
                label_to_nodes[lbl].add(node["id"])
        rows = []
        for group in strategy_data["groups"]:
            _community_id, _count, label, _hex_color = group
            nodes_set = label_to_nodes.get(str(label), set())
            metrics = (
                _subgraph_metrics(nodes_set, graph)
                if nodes_set
                else {
                    "internal_edges": 0,
                    "external_edges": 0,
                    "density": 0.0,
                    "reciprocity": 0.0,
                    "avg_clustering": None,
                    "avg_path_length": None,
                    "diameter": None,
                }
            )
            channels = sorted(
                (
                    {"label": id_to_node[nid].get("label") or nid, "url": id_to_node[nid].get("url") or ""}
                    for nid in nodes_set
                    if nid in id_to_node
                ),
                key=lambda c: c["label"].lower(),
            )
            rows.append({"group": group, "node_count": len(nodes_set), "metrics": metrics, "channels": channels})
        modularity = None
        if label_to_nodes:
            try:
                modularity = nx.community.modularity(graph, label_to_nodes.values())
            except Exception:
                pass
        result["strategies"][strategy_key] = {"modularity": modularity, "rows": rows}
        if status_callback:
            status_callback(strategy_key)
    return result


def _network_summary_rows(summary: dict[str, Any]) -> list[list[Any]]:
    """Return (label, value) rows for all whole-network metrics."""
    path_marker = " *" if not summary["path_on_full"] else ""
    rows: list[list[Any]] = [
        ["Nodes", summary["n"]],
        ["Edges", summary["e"]],
        ["Density", summary["density"]],
        ["Reciprocity", summary["reciprocity"]],
        ["Avg Clustering", summary["avg_clustering"]],
        [f"Avg Path Length{path_marker}", summary["avg_path_length"]],
        [f"Diameter{path_marker}", summary["diameter"]],
        ["WCC count", summary["wcc_count"]],
        ["Largest WCC fraction", summary["wcc_fraction"]],
        ["SCC count", summary["scc_count"]],
        ["Largest SCC fraction", summary["scc_fraction"]],
    ]
    for assort_key, assort_label in [
        ("in_in", "Assortativity in→in"),
        ("in_out", "Assortativity in→out"),
        ("out_in", "Assortativity out→in"),
        ("out_out", "Assortativity out→out"),
    ]:
        rows.append([assort_label, summary.get("assortativity", {}).get(assort_key)])
    if summary.get("mean_burt_constraint") is not None:
        rows.append(["Mean Burt's Constraint", summary["mean_burt_constraint"]])
    if summary.get("network_originality") is not None:
        rows.append(["Content Originality", summary["network_originality"]])
    if summary.get("network_amplification") is not None:
        rows.append(["Amplification Ratio", summary["network_amplification"]])
    for _key, (c_val, c_label) in summary.get("centralizations", {}).items():
        rows.append([f"{c_label} Centralization", c_val])
    return rows


def write_network_metrics_json(
    community_table_data: CommunityTableData,
    strategies: list[str],
    graph_dir: str,
) -> None:
    def _fmt(val: float | None, decimals: int = 4) -> str:
        return "N/A" if val is None else f"{val:.{decimals}f}"

    data_dir = os.path.join(graph_dir, "data")
    summary = community_table_data["network_summary"]
    summary_rows = []
    for label, value in _network_summary_rows(summary):
        if isinstance(value, float):
            display = _fmt(value)
        elif value is None:
            display = "N/A"
        else:
            display = str(value)
        summary_rows.append({"label": label, "value": display})

    modularity_rows = []
    for strategy_key in strategies:
        entry = community_table_data["strategies"].get(strategy_key)
        mod = entry["modularity"] if entry else None
        modularity_rows.append(
            {"strategy": strategy_key.capitalize(), "value": _fmt(mod) if mod is not None else "N/A"}
        )

    payload = {
        "wcc_note_visible": not summary["path_on_full"],
        "summary_rows": summary_rows,
        "modularity_rows": modularity_rows,
    }
    with open(os.path.join(data_dir, "network_metrics.json"), "w") as f:
        f.write(json.dumps(payload))


def write_network_table_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    if seo:
        title = f"{project_title} | Network statistics" if project_title else "Network statistics"
        robots_meta = "index, follow"
    else:
        title = f"{project_title} | Network" if project_title else "Network"
        robots_meta = "noindex, nofollow"

    context = {"title": title, "robots_meta": robots_meta}
    content = render_to_string("network/network_table.html", context)
    with open(output_filename, "w") as f:
        f.write(content)


def write_network_table_xlsx(
    community_table_data: CommunityTableData,
    strategies: list[str],
    output_filename: str,
    project_title: str = "",
) -> None:
    summary = community_table_data["network_summary"]
    wb = openpyxl.Workbook()
    if project_title:
        wb.properties.title = project_title

    ws = wb.active
    ws.title = "Network"
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for label, value in _network_summary_rows(summary):
        ws.append([label, value])
    if not summary["path_on_full"]:
        ws.append([])
        ws.append(["* Computed on the largest weakly connected component (undirected)"])

    ws.append([])
    ws.append(["Strategy", "Modularity"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for strategy_key in strategies:
        entry = community_table_data["strategies"].get(strategy_key)
        ws.append([strategy_key.capitalize(), entry["modularity"] if entry else None])

    wb.save(output_filename)


_COMPARE_RENAMES: dict[str, str] = {
    "graph.html": "graph_2.html",
    "graph3d.html": "graph3d_2.html",
    "channel_table.html": "channel_table_2.html",
    "network_table.html": "network_table_2.html",
    "community_table.html": "community_table_2.html",
    "channel_table.xlsx": "channel_table_2.xlsx",
    "network_table.xlsx": "network_table_2.xlsx",
    "community_table.xlsx": "community_table_2.xlsx",
}

# Ordered so that graph3d.html is replaced before graph.html (avoid partial match)
_HTML_LINK_RENAMES: list[tuple[str, str]] = [
    ("graph3d.html", "graph3d_2.html"),
    ("graph.html", "graph_2.html"),
    ("channel_table.html", "channel_table_2.html"),
    ("network_table.html", "network_table_2.html"),
    ("community_table.html", "community_table_2.html"),
    ("channel_table.xlsx", "channel_table_2.xlsx"),
    ("network_table.xlsx", "network_table_2.xlsx"),
    ("community_table.xlsx", "community_table_2.xlsx"),
]


def _patch_compare_html(content: str) -> str:
    """Patch an HTML file from the compare project: rewrite internal links and inject DATA_DIR."""
    for old, new in _HTML_LINK_RENAMES:
        content = content.replace(old, new)
    injection = '<script>window.DATA_DIR = "data_2/";</script>\n'
    for marker in ('<script src="js/', '<script type="module" src="js/'):
        idx = content.find(marker)
        if idx != -1:
            content = content[:idx] + injection + content[idx:]
            break
    return content


def copy_compare_project(compare_dir: str, graph_dir: str) -> set[str]:
    """Copy files from a compare graph/ directory into graph/, renaming with _2 suffix.

    Returns the set of destination filenames that were actually written.
    """
    copied: set[str] = set()

    # data/ → data_2/
    src_data = os.path.join(compare_dir, "data")
    dst_data = os.path.join(graph_dir, "data_2")
    if os.path.exists(dst_data):
        shutil.rmtree(dst_data)
    if os.path.isdir(src_data):
        shutil.copytree(src_data, dst_data)
        copied.add("data_2")

    for src_name, dst_name in _COMPARE_RENAMES.items():
        src = os.path.join(compare_dir, src_name)
        if not os.path.isfile(src):
            continue
        dst = os.path.join(graph_dir, dst_name)
        if src_name.endswith(".html"):
            with open(src) as f:
                content = f.read()
            with open(dst, "w") as f:
                f.write(_patch_compare_html(content))
        else:
            shutil.copy2(src, dst)
        copied.add(dst_name)

    return copied


def write_network_compare_table_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    if seo:
        title = f"{project_title} | Network comparison" if project_title else "Network comparison"
        robots_meta = "index, follow"
    else:
        title = f"{project_title} | Network comparison" if project_title else "Network comparison"
        robots_meta = "noindex, nofollow"

    context = {"title": title, "robots_meta": robots_meta}
    content = render_to_string("network/network_compare_table.html", context)
    with open(output_filename, "w") as f:
        f.write(content)


def write_community_table_xlsx(
    community_table_data: CommunityTableData,
    strategies: list[str],
    output_filename: str,
    project_title: str = "",
) -> None:
    _HEADERS = [
        "Community",
        "Color",
        "Nodes",
        "Internal Edges",
        "External Edges",
        "Density",
        "Reciprocity",
        "Avg Clustering",
        "Avg Path Length",
        "Diameter",
        "Channels",
    ]

    wb = openpyxl.Workbook()
    if project_title:
        wb.properties.title = project_title
    wb.remove(wb.active)  # no default sheet needed

    # One sheet per strategy
    for strategy_key in strategies:
        strategy_entry = community_table_data["strategies"].get(strategy_key)
        if not strategy_entry:
            continue
        rows = strategy_entry["rows"]
        ws = wb.create_sheet(title=strategy_key.capitalize()[:31])
        ws.append(_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for entry in rows:
            _community_id, _count, label, hex_color = entry["group"]
            hex_color = str(hex_color).lstrip("#")
            metrics = entry["metrics"]
            channels_str = ", ".join(c["label"] for c in entry.get("channels", []))
            ws.append(
                [
                    str(label),
                    f"#{hex_color}",
                    entry["node_count"],
                    metrics["internal_edges"],
                    metrics["external_edges"],
                    metrics["density"],
                    metrics["reciprocity"],
                    metrics["avg_clustering"],
                    metrics["avg_path_length"],
                    metrics["diameter"],
                    channels_str,
                ]
            )
            try:
                fill = PatternFill(start_color=hex_color.upper(), end_color=hex_color.upper(), fill_type="solid")
                ws.cell(row=ws.max_row, column=1).fill = fill
            except Exception:
                pass

    wb.save(output_filename)


def _metric_cell_dict(val: float | int | None, decimals: int, bg: str) -> dict:
    if val is None:
        return {"display": "N/A", "sort_value": "", "style": bg}
    if decimals == 0:
        return {"display": str(int(val)), "sort_value": str(int(val)), "style": bg}
    return {"display": f"{val:.{decimals}f}", "sort_value": str(val), "style": bg}


def write_community_metrics_json(
    community_table_data: CommunityTableData,
    strategies: list[str],
    graph_dir: str,
) -> None:
    data_dir = os.path.join(graph_dir, "data")
    communities_path = os.path.join(data_dir, "communities.json")
    with open(communities_path) as f:
        communities_file = json.load(f)

    for strategy_key in strategies:
        entry = community_table_data["strategies"].get(strategy_key)
        if not entry:
            continue
        rows_out = []
        for row in entry["rows"]:
            _community_id, _count, label, hex_color = row["group"]
            hex_color = str(hex_color)
            if not hex_color.startswith("#"):
                hex_color = f"#{hex_color}"
            rows_out.append(
                {
                    "label": str(label),
                    "hex_color": hex_color,
                    "node_count": row["node_count"],
                    "metrics": row["metrics"],
                    "channels": row.get("channels", []),
                }
            )
        strategy_entry = communities_file["strategies"].get(strategy_key)
        if strategy_entry is not None:
            strategy_entry["rows"] = rows_out

    with open(communities_path, "w") as f:
        f.write(json.dumps(communities_file))


def write_community_table_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
) -> None:
    if seo:
        title = f"{project_title} | Community statistics" if project_title else "Community statistics"
        robots_meta = "index, follow"
    else:
        title = f"{project_title} | Communities" if project_title else "Communities"
        robots_meta = "noindex, nofollow"

    context = {"title": title, "robots_meta": robots_meta}
    content = render_to_string("network/community_table.html", context)
    with open(output_filename, "w") as f:
        f.write(content)


def write_index_html(
    output_filename: str,
    seo: bool = False,
    project_title: str = "",
    include_graph: bool = False,
    include_3d_graph: bool = False,
    include_channel_html: bool = False,
    include_channel_xlsx: bool = False,
    include_network_html: bool = False,
    include_network_xlsx: bool = False,
    include_community_html: bool = False,
    include_community_xlsx: bool = False,
    include_compare_html: bool = False,
    compare_files: set[str] | None = None,
    strategies: list[str] | None = None,
) -> None:
    if seo:
        title = project_title or "Network Analysis"
        robots_meta = "index, follow"
    else:
        title = project_title or "Network Analysis"
        robots_meta = "noindex, nofollow"

    context = {
        "title": title,
        "robots_meta": robots_meta,
        "project_title": project_title,
        "include_graph": include_graph,
        "include_3d_graph": include_3d_graph,
        "include_channel_html": include_channel_html,
        "include_channel_xlsx": include_channel_xlsx,
        "include_network_html": include_network_html,
        "include_network_xlsx": include_network_xlsx,
        "include_community_html": include_community_html,
        "include_community_xlsx": include_community_xlsx,
        "include_compare_html": include_compare_html,
        "compare_files": compare_files or set(),
        "strategies": [s.capitalize() for s in (strategies or [])],
    }
    content = render_to_string("network/index.html", context)
    with open(output_filename, "w") as f:
        f.write(content)
